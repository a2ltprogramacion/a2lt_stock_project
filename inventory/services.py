"""
inventory/services.py
=====================
Motor lógico del sistema A2LT Stock.

Contiene las funciones de negocio críticas que operan sobre la base de datos
de forma transaccional y atómica. Este módulo es la única vía autorizada para
modificar el stock físico en InventarioAlmacen.

REGLA SAGRADA: Ninguna función fuera de este módulo debe alterar directamente
el campo `cantidad_disponible` de InventarioAlmacen.

Funciones principales:
  - registrar_movimiento(...)  → Transacción atómica ENTRADA/SALIDA (Ticket #2)
  - calcular_stock_combo(...)  → Fórmula dinámica de stock de combo (Ticket #2)
  - procesar_salida_combo(...) → Desagregación atómica de componentes (Ticket #2)

MAPA DE SECCIONES (Fase 5 — ref. estructural sin partición de archivo):
  L356   CARGA MASIVA TOLERANTE A FALLOS Y RESOLUCIÓN DE COLISIONES
  L1119  MÓDULO DE VENTAS Y EMISIÓN DE NOTAS DE ENTREGA
  L1299  SINCRONIZACIÓN DE TASAS DE CAMBIO (API)
  L1397  MOTOR DE REVERSO ATÓMICO DE LOTES DE CARGA MASIVA
  L1516  MÓDULO DE MOVIMIENTOS Y AJUSTES MANUALES
  L1623  CONTROL DE COSTOS Y COMPRAS
  L1802  EXPORTACIÓN LÓGICA POR TENANT
  L1865  MÓDULO DE DEVOLUCIONES Y CUARENTENA
  L2013  MÓDULO DE CONTRAPARTIDAS Y REVERSOS (NOTAS Y COMPRAS)

DECISIÓN ADR-21: mantener services.py como módulo único (2085+ líneas) y NO
dividirlo en submódulos mientras todas las pruebas (152+) sigan en verde
y no se identifique una razón funcional fuerte para partirlo. El riesgo de
romper imports circulares (models↔services) supera el beneficio de
legibilidad de un split en este punto. El índice de más arriba funciona
como mapa de navegación.
"""

import logging
from decimal import Decimal

from django.db import transaction

from .models import (
    Articulo,
    Almacen,
    InventarioAlmacen,
    MovimientoKardex,
    NotaEntrega,
    DetalleNotaEntrega,
    ConfiguracionEmpresa,
    Contacto,
)

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN PRINCIPAL: Registrar Movimiento de Kárdex
# ─────────────────────────────────────────────────────────────────────────────

@transaction.atomic
def registrar_movimiento(
    articulo: Articulo,
    almacen: Almacen,
    tipo: str,
    cantidad,
    concepto: str,
    nota_entrega: NotaEntrega = None,
    lote_carga: str = '',
    detalle_adicional: str = '',
    usuario: str = '',
    documento_compra = None,
) -> MovimientoKardex:
    """
    Registra un movimiento de inventario de forma transaccional y atómica.

    Este es el ÚNICO punto de entrada autorizado para modificar el stock físico.
    Si cualquier parte del proceso falla, Django hace rollback de TODA la
    transacción, garantizando que nunca quede el inventario alterado sin su
    registro correspondiente en el Kárdex (Regla Sagrada).

    Args:
        articulo (Articulo): El artículo a mover. Debe ser de tipo 'FISICO'.
        almacen (Almacen):   El almacén origen o destino del movimiento.
        tipo (str):          'ENTRADA' o 'SALIDA'.
        cantidad:            Cantidad a mover (positivo). Se convierte a Decimal.
        concepto (str):      Código de concepto válido de MovimientoKardex.CONCEPTO_CHOICES.
        nota_entrega:        NotaEntrega asociada (opcional, para ventas).
        lote_carga (str):    ID del lote de carga masiva (opcional, para Ticket #3).
        detalle_adicional:   Texto libre de auditoría (opcional).
        usuario (str):       Nombre del usuario que ejecuta el movimiento (opcional).

    Returns:
        MovimientoKardex: El registro creado en el Kárdex.

    Raises:
        ValueError: Si tipo no es 'ENTRADA' o 'SALIDA'.
        ValueError: Si cantidad no es un número positivo.
        ValueError: Si articulo.tipo == 'COMBO' (los combos no tienen stock propio).
        ValueError: En SALIDA, si el stock disponible es insuficiente (anti-negativo).
        Articulo.DoesNotExist / Almacen.DoesNotExist: Si las FK no existen en BD.
    """
    from decimal import Decimal

    # ── Validaciones de entrada ──────────────────────────────────────────────
    tipo = tipo.upper()
    if tipo not in ('ENTRADA', 'SALIDA'):
        raise ValueError(f"Tipo de movimiento inválido: '{tipo}'. Use 'ENTRADA' o 'SALIDA'.")

    cantidad = Decimal(str(cantidad))
    if cantidad <= 0:
        raise ValueError(f"La cantidad debe ser un valor positivo. Recibido: {cantidad}")

    if articulo.tipo == 'COMBO':
        raise ValueError(
            f"El artículo '{articulo.sku}' es de tipo COMBO. "
            "Los combos no tienen stock físico propio. "
            "Use procesar_salida_combo() para desagregar sus componentes."
        )

    # ── Lógica de ENTRADA ────────────────────────────────────────────────────
    from django.db.models import F

    if tipo == 'ENTRADA':
        inv, creado = InventarioAlmacen.objects.get_or_create(
            empresa=articulo.empresa,
            articulo=articulo,
            almacen=almacen,
            defaults={'cantidad_disponible': Decimal('0.00')},
        )
        if not creado:
            # Actualización atómica con F() para prevenir race conditions (ADR-17/C-02)
            InventarioAlmacen.objects.filter(pk=inv.pk).update(
                cantidad_disponible=F('cantidad_disponible') + cantidad
            )
            inv.refresh_from_db()
        else:
            inv.cantidad_disponible = cantidad
            inv.save(update_fields=['cantidad_disponible', 'fecha_ultima_actualizacion'])
        saldo = inv.cantidad_disponible

        logger.info(
            "[KARDEX ENTRADA] %s @ %s | +%s → Saldo: %s | Concepto: %s",
            articulo.sku, almacen.nombre, cantidad, saldo, concepto,
        )

    # ── Lógica de SALIDA ─────────────────────────────────────────────────────
    elif tipo == 'SALIDA':
        # select_for_update() bloquea la fila durante la transacción para
        # prevenir condiciones de carrera en entornos concurrentes.
        try:
            inv = InventarioAlmacen.objects.select_for_update().get(
                articulo=articulo,
                almacen=almacen,
            )
        except InventarioAlmacen.DoesNotExist:
            raise ValueError(
                f"No existe inventario para '{articulo.sku}' en '{almacen.nombre}'. "
                "No se pueden registrar salidas sin existencia previa."
            )

        if inv.cantidad_disponible < cantidad:
            raise ValueError(
                f"Stock insuficiente para '{articulo.nombre}' en '{almacen.nombre}'. "
                f"Disponible: {inv.cantidad_disponible} | Solicitado: {cantidad}. "
                "No se permiten saldos negativos (Regla Sagrada)."
            )

        InventarioAlmacen.objects.filter(pk=inv.pk).update(
            cantidad_disponible=F('cantidad_disponible') - cantidad
        )
        inv.refresh_from_db()
        saldo = inv.cantidad_disponible

        logger.info(
            "[KARDEX SALIDA] %s @ %s | -%s → Saldo: %s | Concepto: %s",
            articulo.sku, almacen.nombre, cantidad, saldo, concepto,
        )

    # ── Creación del registro en el Kárdex ───────────────────────────────────
    movimiento = MovimientoKardex.objects.create(
        empresa=articulo.empresa,
        articulo=articulo,
        almacen=almacen,
        tipo=tipo,
        concepto=concepto,
        cantidad=cantidad,
        saldo_resultante=saldo,
        nota_entrega=nota_entrega,
        documento_compra=documento_compra,
        lote_carga=lote_carga or '',
        detalle_adicional=detalle_adicional or '',
        usuario=usuario or '',
    )

    return movimiento


# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN: Calcular Stock Dinámico de Combo
# ─────────────────────────────────────────────────────────────────────────────

def calcular_stock_combo(combo: Articulo, almacen: Almacen) -> int:
    """
    Calcula el stock disponible de un COMBO de forma dinámica.

    Implementa la fórmula:
        Stock_Combo = min( floor(S(a_i) / q_i) )   para i en receta

    Donde:
        S(a_i) = stock disponible del componente a_i en el almacén dado
        q_i    = cantidad requerida del componente a_i en la receta del combo

    Args:
        combo (Articulo):   Artículo de tipo 'COMBO'.
        almacen (Almacen):  Almacén en el que se evalúa la disponibilidad.

    Returns:
        int: Número entero de combos que se pueden armar (mínimo 0).

    Raises:
        ValueError: Si el artículo no es de tipo 'COMBO'.
    """
    from .models import RecetaCombo  # local para evitar posibles re-imports

    if combo.tipo != 'COMBO':
        raise ValueError(
            f"calcular_stock_combo() solo opera sobre artículos tipo COMBO. "
            f"'{combo.sku}' es de tipo '{combo.tipo}'."
        )

    receta = RecetaCombo.objects.filter(combo=combo).select_related('componente')

    if not receta.exists():
        # Un combo sin receta definida tiene stock 0 (incompleto)
        logger.warning(
            "[COMBO] El combo '%s' no tiene receta definida. Stock = 0.", combo.sku
        )
        return 0

    stocks_posibles = []

    for item in receta:
        componente = item.componente
        cantidad_requerida = item.cantidad_requerida

        # Obtener stock del componente en el almacén especificado
        try:
            inv = InventarioAlmacen.objects.get(
                articulo=componente,
                almacen=almacen,
            )
            stock_componente = inv.cantidad_disponible
        except InventarioAlmacen.DoesNotExist:
            # Si el componente no tiene inventario en este almacén, el combo = 0
            logger.debug(
                "[COMBO] Componente '%s' sin inventario en '%s'. Stock combo = 0.",
                componente.sku, almacen.nombre,
            )
            return 0

        if cantidad_requerida <= 0:
            # Receta inválida: componente con cantidad 0 bloquea el cálculo
            logger.error(
                "[COMBO] Receta inválida: '%s' requiere %s de '%s'.",
                combo.sku, cantidad_requerida, componente.sku,
            )
            return 0

        combos_posibles_con_este = int(stock_componente // cantidad_requerida)
        stocks_posibles.append(combos_posibles_con_este)

    return min(stocks_posibles) if stocks_posibles else 0


# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN: Procesar Salida de Combo (Desagregación Atómica)
# ─────────────────────────────────────────────────────────────────────────────

@transaction.atomic
def procesar_salida_combo(
    combo: Articulo,
    almacen: Almacen,
    cantidad_combos,
    nota_entrega: NotaEntrega = None,
    usuario: str = '',
) -> list:
    """
    Procesa la venta de un COMBO desagregando sus componentes físicos
    de forma atómica en el inventario del almacén indicado.

    El flujo es:
    1. Verifica que el combo tiene stock suficiente (calcular_stock_combo ≥ cantidad_combos).
    2. Para cada componente en la receta, calcula la cantidad a descontar:
           cantidad_a_descontar = cantidad_requerida_en_receta × cantidad_combos
    3. Llama a registrar_movimiento() para cada componente dentro del mismo bloque
       @transaction.atomic. Si CUALQUIER componente falla, TODO el proceso se revierte.

    Args:
        combo (Articulo):       Artículo de tipo 'COMBO' a vender.
        almacen (Almacen):      Almacén del que se despacha.
        cantidad_combos:        Número de combos a despachar.
        nota_entrega:           NotaEntrega asociada (opcional).
        usuario (str):          Usuario que ejecuta la operación.

    Returns:
        list[MovimientoKardex]: Lista de movimientos generados (uno por componente).

    Raises:
        ValueError: Si el artículo no es COMBO.
        ValueError: Si no hay suficiente stock para armar la cantidad de combos pedida.
        ValueError: Si el combo no tiene receta definida.
    """
    from decimal import Decimal
    from .models import RecetaCombo

    if combo.tipo != 'COMBO':
        raise ValueError(
            f"'{combo.sku}' no es un artículo de tipo COMBO."
        )

    cantidad_combos = Decimal(str(cantidad_combos))
    if cantidad_combos <= 0:
        raise ValueError("La cantidad de combos a procesar debe ser mayor a cero.")

    # ── Verificación de disponibilidad previa ────────────────────────────────
    stock_disponible = calcular_stock_combo(combo, almacen)
    if stock_disponible < cantidad_combos:
        raise ValueError(
            f"Stock de combo insuficiente para '{combo.nombre}' en '{almacen.nombre}'. "
            f"Disponible: {stock_disponible} | Solicitado: {cantidad_combos}."
        )

    receta = RecetaCombo.objects.filter(combo=combo).select_related('componente')
    if not receta.exists():
        raise ValueError(
            f"El combo '{combo.sku}' no tiene receta definida. "
            "No se puede procesar la salida."
        )

    movimientos_generados = []
    concepto_base = f'Salida por Venta de Combo [{combo.sku}]'

    logger.info(
        "[COMBO SALIDA] Procesando %s unidades de '%s' en '%s'.",
        cantidad_combos, combo.sku, almacen.nombre,
    )

    for item in receta:
        componente = item.componente
        cantidad_a_descontar = item.cantidad_requerida * cantidad_combos

        movimiento = registrar_movimiento(
            articulo=componente,
            almacen=almacen,
            tipo='SALIDA',
            cantidad=cantidad_a_descontar,
            concepto='VENTA',
            nota_entrega=nota_entrega,
            lote_carga='',
            detalle_adicional=concepto_base,
            usuario=usuario,
        )
        movimientos_generados.append(movimiento)

        logger.info(
            "[COMBO SALIDA] Desagregado: -%s × '%s'",
            cantidad_a_descontar, componente.sku,
        )

    return movimientos_generados


# ═════════════════════════════════════════════════════════════════════════════
# TICKET #3 — CARGA MASIVA TOLERANTE A FALLOS Y RESOLUCIÓN DE COLISIONES
# ═════════════════════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────────────────────────────────────
# HELPER: Validación de Formato de Archivo (ADR-10)
# ─────────────────────────────────────────────────────────────────────────────

def validar_formato_excel(archivo):
    """
    Verifica que el archivo sea .xlsx antes de intentar abrirlo con openpyxl.
    Si el objeto no tiene atributo 'name' (ej: BytesIO en tests), omite la verificación.

    Raises:
        ValueError: Si la extensión no es .xlsx.
    """
    nombre = getattr(archivo, 'name', None)
    if nombre:
        if not nombre.lower().endswith('.xlsx'):
            ext = nombre.rsplit('.', 1)[-1] if '.' in nombre else 'desconocido'
            raise ValueError(
                f"Formato de archivo no soportado: '.{ext}'. "
                "Solo se aceptan archivos Excel en formato .xlsx (Excel 2007 o superior). "
                "Si tiene un archivo .xls, ábralo en Excel y guárdelo como 'Libro Excel (.xlsx)'."
            )


# ─────────────────────────────────────────────────────────────────────────────
# HELPER: Generador de Reporte de Auditoría .txt
# ─────────────────────────────────────────────────────────────────────────────

def _generar_reporte_txt(
    lote_id,
    almacen_nombre,
    filas_procesadas,
    articulos_creados,
    articulos_actualizados,
    filas_error,
    colisiones_pendientes,
    log_errores,
    log_advertencias,
):
    """Genera el texto del reporte de auditoría de la carga masiva."""
    from datetime import datetime

    sep = '=' * 62
    sep_thin = '-' * 62
    ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    lineas = [
        sep,
        '  REPORTE DE CARGA MASIVA — A2LT STOCK',
        sep,
        f'  Lote ID       : {lote_id}',
        f'  Almacén       : {almacen_nombre}',
        f'  Fecha y Hora  : {ahora}',
        sep_thin,
        '  RESUMEN DE PROCESAMIENTO',
        sep_thin,
        f'  Filas procesadas correctamente : {filas_procesadas}',
        f'  Artículos nuevos creados       : {articulos_creados}',
        f'  Artículos actualizados         : {articulos_actualizados}',
        f'  Filas con error aislado        : {filas_error}',
        f'  Colisiones de stock pendientes : {colisiones_pendientes}',
    ]

    if log_errores:
        lineas += [
            sep_thin,
            f'  ERRORES DETECTADOS ({len(log_errores)})',
            sep_thin,
        ]
        lineas.extend(f'  {e}' for e in log_errores)

    if log_advertencias:
        lineas += [
            sep_thin,
            f'  ADVERTENCIAS ({len(log_advertencias)})',
            sep_thin,
        ]
        lineas.extend(f'  {w}' for w in log_advertencias)

    if colisiones_pendientes:
        lineas += [
            sep_thin,
            f'  NOTA: Existen {colisiones_pendientes} colisión(es) de stock pendiente(s) de resolución.',
            '  El inventario físico de esos artículos NO fue modificado.',
            '  Resuélvalas mediante el modal interactivo antes de continuar.',
        ]

    lineas += [sep_thin, '  FIN DEL REPORTE', sep]

    return '\n'.join(lineas)


# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN PRINCIPAL: Procesar Carga Masiva desde Excel (ADR-11, ADR-12, ADR-13)
# ─────────────────────────────────────────────────────────────────────────────

def procesar_carga_masiva(
    archivo_excel,
    almacen_id: int,
    usuario: str = '',
) -> dict:
    """
    Procesa un archivo Excel (.xlsx) de carga masiva de inventario.

    Lógica por fila (ADR-13):
      - SKU nuevo           → Crea Articulo + ENTRADA en Kárdex (si Cantidad > 0)
      - SKU existente + Q=0 → Actualización silenciosa de Nombre/Costo/Precio
      - SKU existente + Q>0 → Colisión: se retorna para resolución interactiva

    Args:
        archivo_excel: Objeto file-like (.xlsx) o BytesIO. Se hace seek(0) interno.
        almacen_id (int): PK del almacén destino por defecto.
        usuario (str): Nombre del usuario que ejecuta la carga.

    Returns:
        dict con claves:
            lote_id (str), almacen_id (int), filas_procesadas (int),
            articulos_creados (int), articulos_actualizados (int),
            filas_error (int), colisiones (list), log_errores (list),
            log_advertencias (list), reporte_txt (str).

    Raises:
        ValueError: Si el almacén no existe o el formato no es .xlsx.
        ValueError: Si openpyxl no puede abrir el archivo (corrupto o vacío).
    """
    import uuid
    import openpyxl
    from decimal import Decimal, InvalidOperation
    from .models import Articulo, Almacen, InventarioAlmacen

    # ── Validaciones previas ─────────────────────────────────────────────────
    validar_formato_excel(archivo_excel)

    try:
        almacen_default = Almacen.objects.get(pk=almacen_id, activo=True)
    except Almacen.DoesNotExist:
        raise ValueError(f"No se encontró un almacén activo con ID {almacen_id}.")

    # ── Seek al inicio (por si el file cursor está al final) ─────────────────
    if hasattr(archivo_excel, 'seek'):
        archivo_excel.seek(0)

    # ── Abrir el workbook ────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(archivo_excel, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(
            f"No se pudo leer el archivo Excel. "
            f"Asegúrese de que es un archivo .xlsx válido y no está dañado. "
            f"Detalle técnico: {exc}"
        )

    ws = wb.active
    lote_id = str(uuid.uuid4())

    # ── Estado acumulador ────────────────────────────────────────────────────
    filas_procesadas = 0
    articulos_creados = 0
    articulos_actualizados = 0
    filas_error = 0
    colisiones = []
    log_errores = []
    log_advertencias = []

    # Pre-carga de nombres existentes para detectar duplicados (warning)
    nombres_existentes = set(
        Articulo.objects.values_list('nombre', flat=True)
    )

    logger.info(
        "[CARGA MASIVA] Iniciando. Lote: %s | Almacén: %s | Usuario: %s",
        lote_id, almacen_default.nombre, usuario or 'anónimo',
    )

    # ── Iteración fila por fila (desde fila 2, saltando cabecera) ────────────
    for fila_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        # Saltar filas completamente vacías
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        try:
            # ── Parseo de columnas ───────────────────────────────────────────
            def celda(idx, default=None):
                """Extrae el valor de la celda por índice (0-based), retorna default si ausente."""
                val = row[idx] if len(row) > idx else None
                if val is None or str(val).strip() == '':
                    return default
                return str(val).strip()

            sku = celda(0, '')
            nombre = celda(1, '')
            costo_raw = celda(2)
            cantidad_raw = celda(3)
            precio_raw = celda(4)
            almacen_nombre_raw = celda(5, '')

            # ── Validaciones obligatorias ────────────────────────────────────
            if not sku:
                raise ValueError("La columna SKU (A) está vacía o nula.")
            if not nombre:
                raise ValueError("La columna Nombre (B) está vacía o nula.")
            if costo_raw is None:
                raise ValueError("La columna Costo (C) es obligatoria y está vacía.")

            # ── Parseo numérico: Costo ───────────────────────────────────────
            try:
                costo = Decimal(costo_raw)
                if costo < 0:
                    raise ValueError("valor negativo")
            except (InvalidOperation, ValueError):
                raise ValueError(
                    f"Costo inválido: '{costo_raw}'. Debe ser un número ≥ 0."
                )

            # ── Parseo numérico: Cantidad (opcional, default 0) ──────────────
            if cantidad_raw is None:
                cantidad = Decimal('0')
            else:
                try:
                    cantidad = Decimal(str(row[3]).strip() if len(row) > 3 and row[3] is not None else '0')
                    if cantidad < 0:
                        raise ValueError("valor negativo")
                except (InvalidOperation, ValueError):
                    raise ValueError(
                        f"Cantidad inválida: '{cantidad_raw}'. Debe ser un número ≥ 0 o estar vacío."
                    )

            # ── Parseo numérico: Precio_Divisa (opcional) ────────────────────
            precio_divisa = None
            if precio_raw is not None:
                try:
                    precio_divisa = Decimal(precio_raw)
                    if precio_divisa < 0:
                        raise ValueError("valor negativo")
                except (InvalidOperation, ValueError):
                    raise ValueError(
                        f"Precio_Divisa inválido: '{precio_raw}'. Debe ser un número ≥ 0 o estar vacío."
                    )

            # ── Resolución de almacén (col F override) ───────────────────────
            almacen = almacen_default
            if almacen_nombre_raw:
                try:
                    almacen = Almacen.objects.get(nombre__iexact=almacen_nombre_raw, activo=True)
                except Almacen.DoesNotExist:
                    raise ValueError(
                        f"Almacén '{almacen_nombre_raw}' no encontrado o inactivo en el sistema."
                    )

            # ── Lógica principal: SKU nuevo vs existente ──────────────────────
            try:
                articulo = Articulo.objects.get(sku=sku)

                # ── SKU EXISTE ────────────────────────────────────────────────
                if cantidad > 0:
                    # COLISIÓN: hay stock que resolver → no tocar inventario aún
                    stock_actual = Decimal('0')
                    try:
                        inv = InventarioAlmacen.objects.get(
                            articulo=articulo, almacen=almacen
                        )
                        stock_actual = inv.cantidad_disponible
                    except InventarioAlmacen.DoesNotExist:
                        pass

                    colisiones.append({
                        'fila': fila_num,
                        'sku': sku,
                        'nombre': articulo.nombre,
                        'stock_actual': str(stock_actual),
                        'cantidad_excel': str(cantidad),
                        'almacen_id': almacen.pk,
                        'almacen_nombre': almacen.nombre,
                        'costo': str(costo),
                        'precio_divisa': str(precio_divisa) if precio_divisa is not None else '',
                        'nombre_excel': nombre,
                        'lote_id': lote_id,
                    })
                    logger.info(
                        "[CARGA MASIVA] Colisión detectada: SKU '%s' @ '%s'. "
                        "Stock actual: %s | Cantidad Excel: %s",
                        sku, almacen.nombre, stock_actual, cantidad,
                    )
                else:
                    # ACTUALIZACIÓN SILENCIOSA: sin cantidad → solo actualiza campos base
                    update_fields = ['nombre', 'costo', 'fecha_actualizacion']
                    articulo.nombre = nombre
                    articulo.costo = costo
                    if precio_divisa is not None:
                        articulo.precio_divisa = precio_divisa
                        update_fields.append('precio_divisa')
                    articulo.save(update_fields=update_fields)
                    articulos_actualizados += 1
                    logger.debug(
                        "[CARGA MASIVA] Actualización silenciosa: SKU '%s'.", sku
                    )

            except Articulo.DoesNotExist:
                # ── SKU NUEVO: crear artículo ─────────────────────────────────
                if nombre in nombres_existentes:
                    advertencia = (
                        f"Fila {fila_num} [{sku}] [Advertencia]: "
                        f"Se creó el SKU '{sku}' con un nombre idéntico a un artículo "
                        f"ya existente en el catálogo ('{nombre}')."
                    )
                    log_advertencias.append(advertencia)
                    logger.warning(advertencia)

                articulo = Articulo.objects.create(
                    empresa=almacen.empresa,
                    sku=sku,
                    nombre=nombre,
                    tipo='FISICO',
                    costo=costo,
                    precio_divisa=precio_divisa if precio_divisa is not None else Decimal('0'),
                )
                nombres_existentes.add(nombre)
                articulos_creados += 1

                # Si viene con cantidad, registrar la entrada inicial en el Kárdex
                if cantidad > 0:
                    registrar_movimiento(
                        articulo=articulo,
                        almacen=almacen,
                        tipo='ENTRADA',
                        cantidad=cantidad,
                        concepto='CARGA_MASIVA_SUMA',
                        lote_carga=lote_id,
                        detalle_adicional=(
                            f'Entrada inicial por Carga Masiva — Artículo nuevo. '
                            f'Lote: {lote_id}'
                        ),
                        usuario=usuario,
                    )

                logger.info(
                    "[CARGA MASIVA] Artículo nuevo creado: SKU '%s' | Cantidad: %s",
                    sku, cantidad,
                )

            filas_procesadas += 1

        except Exception as exc:
            sku_log = str(row[0]).strip() if row and row[0] is not None else 'N/A'
            error_msg = f"Fila {fila_num} [{sku_log}]: {exc}"
            log_errores.append(error_msg)
            filas_error += 1
            logger.warning("[CARGA MASIVA] Error aislado en %s", error_msg)

    try:
        wb.close()
    except Exception:
        pass

    reporte_txt = _generar_reporte_txt(
        lote_id=lote_id,
        almacen_nombre=almacen_default.nombre,
        filas_procesadas=filas_procesadas,
        articulos_creados=articulos_creados,
        articulos_actualizados=articulos_actualizados,
        filas_error=filas_error,
        colisiones_pendientes=len(colisiones),
        log_errores=log_errores,
        log_advertencias=log_advertencias,
    )

    logger.info(
        "[CARGA MASIVA] Finalizado. Lote: %s | OK: %s | Errores: %s | Colisiones: %s",
        lote_id, filas_procesadas, filas_error, len(colisiones),
    )

    return {
        'lote_id': lote_id,
        'almacen_id': almacen_id,
        'filas_procesadas': filas_procesadas,
        'articulos_creados': articulos_creados,
        'articulos_actualizados': articulos_actualizados,
        'filas_error': filas_error,
        'colisiones': colisiones,
        'log_errores': log_errores,
        'log_advertencias': log_advertencias,
        'reporte_txt': reporte_txt,
    }


# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN: Procesar Carga Masiva desde Excel — Atómica y Estricta (Ticket #27)
# ─────────────────────────────────────────────────────────────────────────────

def procesar_carga_masiva_excel(file_io, empresa_id, usuario=''):
    """
    Procesa un archivo Excel de carga masiva de forma estrictamente atómica.

    Reglas (Ticket #27):
      - Valida cabeceras exactas: SKU, Nombre, Costo, Cantidad, Precio_Divisa, Almacen.
      - Valida que cada almacén pertenezca al Tenant activo.
      - Casteo seguro a Decimal. Valores negativos o no numéricos → ValueError.
      - SKU existente → actualiza campos base + registrar_movimiento ENTRADA.
      - SKU nuevo → crea Articulo + registrar_movimiento ENTRADA.
      - Envuelto en transaction.atomic(): cualquier error → rollback total.

    Args:
        file_io: BytesIO o file-like del archivo .xlsx.
        empresa_id (int): PK de la empresa (Tenant) activa.
        usuario (str): Nombre del operador que ejecuta la carga.

    Returns:
        dict con claves: lote_id, filas_procesadas, articulos_creados, kardex_entradas.

    Raises:
        ValueError: Cabeceras incorrectas, almacén ajeno al Tenant, valores inválidos.
    """
    import uuid
    import openpyxl
    from decimal import Decimal, InvalidOperation
    from .models import Articulo, Almacen, Empresa

    validar_formato_excel(file_io)

    if hasattr(file_io, 'seek'):
        file_io.seek(0)

    try:
        wb = openpyxl.load_workbook(file_io, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"No se pudo leer el archivo Excel: {exc}")

    ws = wb.active
    lote_id = str(uuid.uuid4())

    # ── Validación de cabeceras ──────────────────────────────────────────────
    EXPECTED_HEADERS = ['SKU', 'Nombre', 'Costo', 'Cantidad', 'Precio_Divisa', 'Almacen']
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else '' for h in first_row]
    if not first_row or headers != EXPECTED_HEADERS:
        try:
            wb.close()
        except Exception:
            pass
        raise ValueError(
            f"Las cabeceras del archivo no coinciden con la plantilla esperada.\n"
            f"Esperado: {EXPECTED_HEADERS}\n"
            f"Recibido:  {headers}"
        )

    empresa = Empresa.objects.get(pk=empresa_id)

    # ── Pre-carga de almacenes del Tenant ────────────────────────────────────
    almacenes_tenant = {
        a.nombre.lower(): a
        for a in Almacen.objects.filter(empresa_id=empresa_id, activo=True)
    }
    if not almacenes_tenant:
        try:
            wb.close()
        except Exception:
            pass
        raise ValueError(f"No hay almacenes activos registrados para la empresa '{empresa.nombre}'.")

    filas_procesadas = 0
    articulos_creados = 0
    kardex_entradas = 0

    try:
        with transaction.atomic():
            for fila_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

                # Saltar filas completamente vacías
                if all(v is None or str(v).strip() == '' for v in row):
                    continue

                # ── Parseo de columnas ───────────────────────────────────────
                sku = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
                nombre = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                costo_raw = row[2] if len(row) > 2 else None
                cantidad_raw = row[3] if len(row) > 3 else None
                precio_raw = row[4] if len(row) > 4 else None
                almacen_nombre_raw = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ''

                # ── Validaciones obligatorias ────────────────────────────────
                if not sku:
                    raise ValueError(f"Fila {fila_num}: SKU vacío.")
                if not nombre:
                    raise ValueError(f"Fila {fila_num} [{sku}]: Nombre vacío.")
                if costo_raw is None or str(costo_raw).strip() == '':
                    raise ValueError(f"Fila {fila_num} [{sku}]: Costo obligatorio ausente.")

                # ── Casteo seguro a Decimal ──────────────────────────────────
                try:
                    costo = Decimal(str(costo_raw))
                    if costo < 0:
                        raise ValueError("negativo")
                except (InvalidOperation, ValueError):
                    raise ValueError(
                        f"Fila {fila_num} [{sku}]: Costo inválido '{costo_raw}'. Debe ser un número ≥ 0."
                    )

                if cantidad_raw is None or str(cantidad_raw).strip() == '':
                    cantidad = Decimal('0')
                else:
                    try:
                        cantidad = Decimal(str(cantidad_raw))
                        if cantidad < 0:
                            raise ValueError("negativo")
                    except (InvalidOperation, ValueError):
                        raise ValueError(
                            f"Fila {fila_num} [{sku}]: Cantidad inválida '{cantidad_raw}'."
                        )

                if precio_raw is not None and str(precio_raw).strip() != '':
                    try:
                        precio_divisa = Decimal(str(precio_raw))
                        if precio_divisa < 0:
                            raise ValueError("negativo")
                    except (InvalidOperation, ValueError):
                        raise ValueError(
                            f"Fila {fila_num} [{sku}]: Precio_Divisa inválido '{precio_raw}'."
                        )
                else:
                    precio_divisa = None

                # ── Validación de Almacén (pertenece al Tenant) ──────────────
                if not almacen_nombre_raw:
                    raise ValueError(f"Fila {fila_num} [{sku}]: Almacén no especificado.")

                almacen_obj = almacenes_tenant.get(almacen_nombre_raw.lower())
                if not almacen_obj:
                    raise ValueError(
                        f"Fila {fila_num} [{sku}]: Almacén '{almacen_nombre_raw}' "
                        f"no encontrado o no pertenece a la empresa '{empresa.nombre}'."
                    )

                # ── Lógica principal: SKU existente o nuevo ──────────────────
                try:
                    articulo = Articulo.objects.get(sku=sku, empresa_id=empresa_id)
                    # SKU EXISTE: actualizar campos base
                    articulo.nombre = nombre
                    articulo.costo = costo
                    update_fields = ['nombre', 'costo', 'fecha_actualizacion']
                    if precio_divisa is not None:
                        articulo.precio_divisa = precio_divisa
                        update_fields.append('precio_divisa')
                    articulo.save(update_fields=update_fields)
                except Articulo.DoesNotExist:
                    # SKU NUEVO: crear artículo
                    articulo = Articulo.objects.create(
                        empresa=empresa,
                        sku=sku,
                        nombre=nombre,
                        tipo='FISICO',
                        categoria='OTROS',
                        costo=costo,
                        precio_divisa=precio_divisa if precio_divisa is not None else Decimal('0'),
                    )
                    articulos_creados += 1

                # ── Registrar movimiento de ENTRADA en Kárdex ────────────────
                if cantidad > 0:
                    registrar_movimiento(
                        articulo=articulo,
                        almacen=almacen_obj,
                        tipo='ENTRADA',
                        cantidad=cantidad,
                        concepto='CARGA_MASIVA_SUMA',
                        lote_carga=lote_id,
                        detalle_adicional=f'Entrada por Carga Masiva Excel — Lote {lote_id}',
                        usuario=usuario,
                    )
                    kardex_entradas += 1

                filas_procesadas += 1

    except Exception:
        try:
            wb.close()
        except Exception:
            pass
        raise

    try:
        wb.close()
    except Exception:
        pass

    logger.info(
        "[CARGA MASIVA EXCEL] Finalizado. Lote: %s | Filas: %s | Creados: %s | Kárdex: %s",
        lote_id, filas_procesadas, articulos_creados, kardex_entradas,
    )

    return {
        'lote_id': lote_id,
        'filas_procesadas': filas_procesadas,
        'articulos_creados': articulos_creados,
        'kardex_entradas': kardex_entradas,
    }


# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN: Resolver Colisión de SKU (Los 3 Botones del Modal)
# ─────────────────────────────────────────────────────────────────────────────

@transaction.atomic
def resolver_colision(
    sku: str,
    almacen_id: int,
    decision: str,
    cantidad_excel,
    lote_id: str,
    usuario: str = '',
    costo=None,
    precio_divisa=None,
    nombre_excel: str = '',
) -> dict:
    """
    Aplica la decisión del usuario sobre un SKU en colisión de forma atómica.

    Decisiones:
        'SUMAR'     → ENTRADA por cantidad_excel sumándose al stock actual.
        'SUSTITUIR' → SALIDA del stock actual (→ 0) + ENTRADA del valor Excel.
        'CANCELAR'  → No modifica el inventario. Solo retorna confirmación.

    Args:
        sku (str):              SKU del artículo en conflicto.
        almacen_id (int):       PK del almacén donde resolver.
        decision (str):         'SUMAR' | 'SUSTITUIR' | 'CANCELAR'.
        cantidad_excel:         Cantidad del archivo Excel (Decimal o str numérico).
        lote_id (str):          UUID del lote de carga masiva.
        usuario (str):          Nombre del operador.
        costo:                  Nuevo costo a aplicar al artículo (opcional).
        precio_divisa:          Nuevo precio a aplicar (opcional).
        nombre_excel (str):     Nombre del artículo proveniente del Excel (opcional).

    Returns:
        dict con claves: decision, sku, movimientos (list de IDs), mensaje.

    Raises:
        ValueError: Si la decisión no es válida.
        Articulo.DoesNotExist: Si el SKU no existe en la BD.
        Almacen.DoesNotExist: Si el almacén no existe.
    """
    from decimal import Decimal
    from .models import Articulo, Almacen, InventarioAlmacen

    decision = decision.upper().strip()
    if decision not in ('SUMAR', 'SUSTITUIR', 'CANCELAR'):
        raise ValueError(
            f"Decisión inválida: '{decision}'. "
            "Las opciones válidas son: SUMAR, SUSTITUIR, CANCELAR."
        )

    if decision == 'CANCELAR':
        logger.info(
            "[COLISIÓN CANCELAR] SKU '%s' | Lote: %s — Sin cambios.",
            sku, lote_id,
        )
        return {
            'decision': 'CANCELAR',
            'sku': sku,
            'movimientos': [],
            'mensaje': f"El artículo '{sku}' fue ignorado. El stock actual no fue modificado.",
        }

    articulo = Articulo.objects.get(sku=sku)
    almacen = Almacen.objects.get(pk=almacen_id)
    cantidad_excel = Decimal(str(cantidad_excel))

    # Aplicar actualizaciones de campos base del artículo si se proveen
    campos_actualizar = []
    if nombre_excel and nombre_excel != articulo.nombre:
        articulo.nombre = nombre_excel
        campos_actualizar.append('nombre')
    if costo is not None:
        articulo.costo = Decimal(str(costo))
        campos_actualizar.append('costo')
    if precio_divisa is not None and str(precio_divisa) != '':
        articulo.precio_divisa = Decimal(str(precio_divisa))
        campos_actualizar.append('precio_divisa')
    if campos_actualizar:
        campos_actualizar.append('fecha_actualizacion')
        articulo.save(update_fields=campos_actualizar)

    movimientos_ids = []

    if decision == 'SUMAR':
        movimiento = registrar_movimiento(
            articulo=articulo,
            almacen=almacen,
            tipo='ENTRADA',
            cantidad=cantidad_excel,
            concepto='CARGA_MASIVA_SUMA',
            lote_carga=lote_id,
            detalle_adicional=f'Entrada por Carga Masiva (Suma) — Lote {lote_id}',
            usuario=usuario,
        )
        movimientos_ids.append(movimiento.pk)
        mensaje = (
            f"SUMAR: Se añadieron {cantidad_excel} unidades a '{articulo.nombre}' "
            f"en '{almacen.nombre}'."
        )
        logger.info("[COLISIÓN SUMAR] SKU '%s' | +%s | Lote: %s", sku, cantidad_excel, lote_id)

    elif decision == 'SUSTITUIR':
        # Paso 1: Vaciar stock actual con SALIDA (si hay stock > 0)
        stock_actual = Decimal('0')
        try:
            inv = InventarioAlmacen.objects.select_for_update().get(
                articulo=articulo, almacen=almacen
            )
            stock_actual = inv.cantidad_disponible
        except InventarioAlmacen.DoesNotExist:
            pass

        if stock_actual > 0:
            m_salida = registrar_movimiento(
                articulo=articulo,
                almacen=almacen,
                tipo='SALIDA',
                cantidad=stock_actual,
                concepto='CARGA_MASIVA_SUSTITUCION_SALIDA',
                lote_carga=lote_id,
                detalle_adicional=(
                    f'Ajuste de Salida por Sustitución — Carga Masiva Lote {lote_id}'
                ),
                usuario=usuario,
            )
            movimientos_ids.append(m_salida.pk)

        # Paso 2: Asentar el nuevo valor del Excel con ENTRADA
        m_entrada = registrar_movimiento(
            articulo=articulo,
            almacen=almacen,
            tipo='ENTRADA',
            cantidad=cantidad_excel,
            concepto='CARGA_MASIVA_SUSTITUCION_ENTRADA',
            lote_carga=lote_id,
            detalle_adicional=(
                f'Entrada por Sustitución — Carga Masiva Lote {lote_id}'
            ),
            usuario=usuario,
        )
        movimientos_ids.append(m_entrada.pk)
        mensaje = (
            f"SUSTITUIR: Stock de '{articulo.nombre}' en '{almacen.nombre}' "
            f"reemplazado: {stock_actual} → {cantidad_excel} unidades. "
            f"Registrados {len(movimientos_ids)} movimientos en el Kárdex."
        )
        logger.info(
            "[COLISIÓN SUSTITUIR] SKU '%s' | %s→%s | Lote: %s",
            sku, stock_actual, cantidad_excel, lote_id,
        )

    return {
        'decision': decision,
        'sku': sku,
        'movimientos': movimientos_ids,
        'mensaje': mensaje,
    }


# ═══════════════════════════════════════════════════════════════════════
# TICKET #5: MÓDULO DE VENTAS Y EMISIÓN DE NOTAS DE ENTREGA
# ═══════════════════════════════════════════════════════════════════════

@transaction.atomic
def procesar_venta(cliente_id=None, lista_items=None, almacen_id=None, usuario='',
                    observaciones='', empresa_id=None,
                    tipo_documento='NOTA_ENTREGA', numero_factura='',
                    descuento_global=Decimal('0')) -> NotaEntrega:
    """
    Procesa una venta generando una Nota de Entrega o Factura con inmutabilidad de precios.

    Parámetros (Fase N2):
      - tipo_documento: 'NOTA_ENTREGA' (default) o 'FACTURA'.
      - numero_factura: string manual. También si tipo_documento=FACTURA. Único por empresa.
      - descuento_global: Decimal. % descuento global sobre el documento.

    IVA: Se calcula siempre de forma individual por artículo (iva_porcentaje snapshot).
    No existe un switch global de IVA — cada item declara su propio impuesto.

    Flujo atómico:
    1. Valida blindaje multi-pestaña (empresa_id del payload vs contexto activo).
    2. Lee el snapshot cambiario de ConfiguracionEmpresa (tasa_bcv + factor + tasa_mercado).
    3. Valida existencias para los artículos físicos (los combos se validan vía sus componentes).
    4. Registra la salida en el Kárdex para cada artículo físico o componentes de combo.
    5. Crea la NotaEntrega (con tipo_documento, numero_factura, descuento_global,
       tasa_mercado_aplicada snapshot) y sus Detalles (4 precios snapshot + IVA snapshot).
    6. Calcula iva_total = Σ iva_bs de cada detalle y lo persiste.
    """
    from .managers import get_current_empresa
    from decimal import Decimal
    # Validación perimetral Multi-Tenant contra contaminación de sesión cruzada
    _ctx_empresa = get_current_empresa()

    if _ctx_empresa is None:
        raise ValueError("Seguridad Contable: No se detectó un contexto de Tenant activo para esta transacción.")

    if empresa_id is None:
        empresa_id = _ctx_empresa

    if not empresa_id:
        raise ValueError("Seguridad Contable: El identificador de la empresa emisora es obligatorio en el payload.")

    try:
        empresa_id_int = int(empresa_id)
        ctx_int = int(_ctx_empresa)
    except (ValueError, TypeError):
        raise ValueError("Seguridad Contable: El identificador del Tenant en el payload es inválido o ha sido alterado.")

    if empresa_id_int != ctx_int:
        raise ValueError("El contexto de la empresa ha cambiado en otra pestaña. Petición abortada por seguridad contable.")

    # Validacion multi-tenant: el almacen debe pertenecer a la empresa activa.
    # Ademas capturamos el almacen global para tener su empresa_id y asi
    # resolver ConfiguracionEmpresa con el tenant correcto del contexto
    # (que debe coincidir forzosamente con el almacen).
    try:
        almacen = Almacen.global_objects.get(pk=almacen_id, empresa_id=empresa_id_int)
    except Almacen.DoesNotExist:
        raise ValueError(
            f"El almacen {almacen_id} no pertenece a la empresa activa "
            f"o no existe."
        )

    # 1. Configuración Cambiaria (Blindaje y prevención de ZeroDivisionError)
    config = ConfiguracionEmpresa.objects.get(empresa_id=almacen.empresa_id)
    if config.tasa_bcv <= 0:
        raise ValueError("Error de Configuración: La Tasa BCV debe ser mayor a 0 para facturar.")

    tasa_aplicada = config.tasa_bcv
    factor_aplicado = config.factor_cobertura
    tasa_mercado_aplicada = config.tasa_mercado

    # N2: Validación de tipo_documento y numero_factura
    if tipo_documento not in ('NOTA_ENTREGA', 'FACTURA'):
        raise ValueError(
            f"Tipo de documento inválido: '{tipo_documento}'. "
            "Debe ser 'NOTA_ENTREGA' o 'FACTURA'."
        )
    if tipo_documento == 'FACTURA':
        if not numero_factura or not str(numero_factura).strip():
            raise ValueError(
                "Seguridad Contable: numero_factura es obligatorio cuando "
                "tipo_documento='FACTURA'."
            )
        numero_factura = str(numero_factura).strip()
        # Validar unicidad previa (la constraint DB también protege, pero damos
        # un mensaje claro antes de llegar al IntegrityError).
        if NotaEntrega.global_objects.filter(
            empresa_id=almacen.empresa_id,
            numero_factura=numero_factura,
        ).exists():
            raise ValueError(
                f"El número de factura '{numero_factura}' ya existe "
                "para esta empresa. No se pueden duplicar facturas."
            )
    else:
        # NOTA_ENTREGA: número_factura debe ir vacío.
        numero_factura = ''

    # N2: Validación descuento_global (0-100)
    try:
        descuento_global = Decimal(str(descuento_global))
    except (ValueError, TypeError):
        raise ValueError("descuento_global debe ser un número decimal entre 0 y 100.")
    if not (Decimal('0') <= descuento_global <= Decimal('100')):
        raise ValueError("descuento_global debe estar entre 0 y 100.")

    # 2. Resolución del Cliente (el almacen ya esta validado arriba)
    
    if cliente_id:
        cliente = Contacto.objects.get(pk=cliente_id)
    else:
        # Cliente genérico por defecto
        cliente, _ = Contacto.objects.get_or_create(
            empresa=almacen.empresa,
            tipo='CLIENTE',
            nombre='Cliente Genérico',
            defaults={'identificacion': 'V-00000000'}
        )

    # 3. Creación de la Cabecera (Nota de Entrega o Factura)
    nota_entrega = NotaEntrega.objects.create(
        empresa=almacen.empresa,
        cliente=cliente,
        almacen=almacen,
        tasa_bcv_aplicada=tasa_aplicada,
        factor_cobertura_aplicado=factor_aplicado,
        tasa_mercado_aplicada=tasa_mercado_aplicada,
        observaciones=observaciones,
        tipo_documento=tipo_documento,
        numero_factura=numero_factura,
        descuento_global=descuento_global,
    )

    # 4. Procesamiento de Ítems
    for item in lista_items:
        articulo = Articulo.objects.get(sku=item['articulo_sku'])
        cantidad = Decimal(str(item['cantidad']))
        # Compatibilidad: acepta precio_unitario_usd (legacy) o precio_base (nuevo).
        precio_base = Decimal(str(
            item.get('precio_base', item.get('precio_unitario_usd'))
        ))
        # Descuento individual opcional (default 0)
        descuento_aplicado = Decimal(str(item.get('descuento_aplicado', '0')))
        # IVA por línea opcional (default 0, sobrescribe Articulo.iva_porcentaje
        # si la UI lo envía; NO aceptamos valor fuera de 0-100).
        iva_porcentaje_item = item.get('iva_porcentaje', None)
        if iva_porcentaje_item is not None:
            iva_porcentaje_item = Decimal(str(iva_porcentaje_item))
            if not (Decimal('0') <= iva_porcentaje_item <= Decimal('100')):
                raise ValueError(
                    f"iva_porcentaje para {articulo.sku} debe estar entre 0 y 100."
                )

        # Validación estricta de Stock
        stock_disp = articulo.get_stock_disponible(almacen)
        if stock_disp < cantidad:
            raise ValueError(
                f"Stock insuficiente para '{articulo.nombre}'. "
                f"Requerido: {cantidad}, Disponible: {stock_disp} en '{almacen.nombre}'."
            )

        # Descuento atómico del Kárdex
        if articulo.tipo == 'FISICO':
            registrar_movimiento(
                articulo=articulo,
                almacen=almacen,
                tipo='SALIDA',
                cantidad=cantidad,
                concepto='VENTA',
                nota_entrega=nota_entrega,
                usuario=usuario
            )
        elif articulo.tipo == 'COMBO':
            procesar_salida_combo(
                combo=articulo,
                almacen=almacen,
                cantidad_combos=cantidad,
                nota_entrega=nota_entrega,
                usuario=usuario
            )

        # Inmutabilidad Financiera: 4 snapshots de precios al momento de la venta
        precio_ajustado = (precio_base * factor_aplicado).quantize(Decimal('0.0001'))
        precio_directo_bcv = (precio_base * tasa_aplicada).quantize(Decimal('0.01'))
        precio_ajustado_bcv = (precio_ajustado * tasa_aplicada).quantize(Decimal('0.01'))

        # IVA snapshot — preferir el del item si la UI lo manda (1.1.2 O2);
        # caer al del Articulo si no viene (compat hacia atrás).
        if iva_porcentaje_item is not None:
            iva_porcentaje = iva_porcentaje_item
        else:
            iva_porcentaje = articulo.iva_porcentaje

        # Creación de Detalle (Línea de factura) — ADR-18: 4 precios snapshot + IVA
        detalle_nota = DetalleNotaEntrega.objects.create(
            nota_entrega=nota_entrega,
            articulo=articulo,
            almacen=almacen,
            cantidad=cantidad,
            precio_base=precio_base,
            precio_ajustado=precio_ajustado,
            precio_directo_bcv=precio_directo_bcv,
            precio_ajustado_bcv=precio_ajustado_bcv,
            costo_unitario_snapshot=articulo.costo,
            descuento_aplicado=descuento_aplicado,
            iva_porcentaje=iva_porcentaje,
        )

        # Validación estricta y quema atómica de Seriales (Ticket #14-SAAS)
        if articulo.usa_serial:
            from inventory.models import SerialArticulo
            seriales_ids = item.get('seriales', [])
            if len(seriales_ids) != int(cantidad):
                raise ValueError(
                    f"El artículo '{articulo.nombre}' requiere exactamente {int(cantidad)} seriales, "
                    f"pero se enviaron {len(seriales_ids)}."
                )

            # Bloqueo pesimista: select_for_update garantiza que nadie más pueda tocar estos seriales
            # simultáneamente en otra transacción.
            seriales_db = list(SerialArticulo.objects.select_for_update().filter(
                serial__in=seriales_ids,
                articulo=articulo,
                empresa=almacen.empresa,
                almacen=almacen
            ))

            if len(seriales_db) != len(seriales_ids):
                encontrados = [s.serial for s in seriales_db]
                faltantes = set(seriales_ids) - set(encontrados)
                raise ValueError(
                    f"Los siguientes seriales no se encontraron o no pertenecen a '{almacen.nombre}': {', '.join(faltantes)}"
                )

            for s in seriales_db:
                if s.estado != 'DISPONIBLE':
                    raise ValueError(
                        f"El serial '{s.serial}' del artículo '{articulo.nombre}' ya no está DISPONIBLE "
                        f"(Posible venta concurrente)."
                    )
                # Mutación de estado y amarre
                s.estado = 'VENDIDO'
                s.detalle_nota = detalle_nota
            
            SerialArticulo.objects.bulk_update(seriales_db, ['estado', 'detalle_nota'])

    # N2: Cálculo de iva_total (siempre, desde snapshot individual de cada detalle)
    iva_total_calc = Decimal('0')
    for det in nota_entrega.detalles.all():
        iva_total_calc += det.iva_bs
    nota_entrega.iva_total = iva_total_calc.quantize(Decimal('0.0001'))
    # iva_check refleja si el documento lleva IVA (algún item con iva_porcentaje > 0)
    nota_entrega.iva_check = iva_total_calc > 0
    nota_entrega.save(update_fields=['iva_total', 'iva_check'])

    # El objeto en memoria ya tiene el valor correcto de iva_total
    logger.info(
        "Venta procesada exitosamente. Nota %s | %s ítems | Tasa: %s | Tipo: %s",
        nota_entrega.numero, len(lista_items), tasa_aplicada, tipo_documento
    )

    return nota_entrega


# ─────────────────────────────────────────────────────────────────────────────
# 5. SINCRONIZACIÓN DE TASAS DE CAMBIO (API)
# ─────────────────────────────────────────────────────────────────────────────

def sincronizar_tasa_cambio() -> dict:
    """
    Se conecta a la API configurada para obtener la tasa de mercado,
    calcula el nuevo factor de cobertura y registra la auditoría.
    """
    import requests
    from decimal import Decimal, InvalidOperation
    from .models import ConfiguracionEmpresa, AuditoriaTasa

    from .managers import get_current_empresa
    _empresa_id = get_current_empresa()
    if _empresa_id is None:
        return {'ok': False, 'error': 'No se pudo determinar la empresa activa para sincronizar tasas.'}
    config = ConfiguracionEmpresa.global_objects.get(empresa_id=_empresa_id)

    if not config.api_url:
        return {'ok': False, 'error': "La URL de la API no está configurada en el sistema."}

    # 1. Petición HTTP
    try:
        response = requests.request(
            method=config.http_method.upper(),
            url=config.api_url,
            timeout=5
        )
        response.raise_for_status()
        data = response.json()
    except requests.exceptions.Timeout:
        logger.error("[API] Timeout al conectar con la API de tasas de cambio.")
        return {'ok': False, 'error': "Tiempo de espera agotado al conectar con el servidor."}
    except requests.exceptions.RequestException as e:
        logger.error("[API] Error de conexión: %s", e)
        return {'ok': False, 'error': f"Error de red o servidor: {str(e)}"}
    except ValueError:
        logger.error("[API] La respuesta no es un JSON válido.")
        return {'ok': False, 'error': "La respuesta del servidor no tiene formato JSON."}

    # 2. Extractor Seguro (Notación de puntos, cero eval())
    selector = config.response_selector or ""
    keys = [k for k in selector.split('.') if k]
    
    current_val = data
    for key in keys:
        if isinstance(current_val, dict) and key in current_val:
            current_val = current_val[key]
        elif isinstance(current_val, list) and key.isdigit() and int(key) < len(current_val):
            current_val = current_val[int(key)]
        else:
            logger.error("[API] El selector '%s' falló en la llave '%s'.", selector, key)
            return {'ok': False, 'error': f"No se pudo encontrar la ruta '{selector}' en la respuesta."}

    try:
        tasa_mercado_obtenida = Decimal(str(current_val))
    except (ValueError, TypeError, InvalidOperation):
        logger.error("[API] El valor obtenido no es numérico: %s", current_val)
        return {'ok': False, 'error': "El valor devuelto por la API no es un número válido."}

    if tasa_mercado_obtenida <= 0:
        return {'ok': False, 'error': "La tasa obtenida debe ser mayor a 0."}

    # 3. Asentamiento Transaccional
    with transaction.atomic():
        # Bloqueo optimista del registro de configuración
        config_db = ConfiguracionEmpresa.objects.select_for_update().get(pk=config.pk)
        
        tasa_bcv_actual = config_db.tasa_bcv
        if tasa_bcv_actual <= 0:
            return {'ok': False, 'error': "La tasa BCV base no está configurada o es 0."}

        # Cálculo del factor de cobertura: T_mercado / T_bcv
        factor_nuevo = tasa_mercado_obtenida / tasa_bcv_actual

        config_db.tasa_mercado = tasa_mercado_obtenida
        config_db.factor_cobertura = factor_nuevo
        config_db.save()

        # Registro inalterable de auditoría
        AuditoriaTasa.objects.create(
            empresa=config_db.empresa,
            tasa_bcv=tasa_bcv_actual,
            tasa_mercado=tasa_mercado_obtenida,
            factor_cobertura=factor_nuevo,
            fuente='API'
        )
    
    logger.info("[API] Tasa sincronizada exitosamente: %s (Factor: %s)", tasa_mercado_obtenida, factor_nuevo)
    return {
        'ok': True,
        'tasa_mercado': float(tasa_mercado_obtenida),
        'factor_cobertura': float(factor_nuevo)
    }


# ─────────────────────────────────────────────────────────────────────────────
# 7. MOTOR DE REVERSO ATÓMICO DE LOTES DE CARGA MASIVA
# ─────────────────────────────────────────────────────────────────────────────

def revertir_carga_masiva(lote_id: str, usuario: str = '') -> dict:
    """
    Deshace atómicamente una carga masiva completa.
    Verifica que el inventario no haya sido alterado con salidas
    (ventas, ajustes, combos) posteriores a la carga del lote.
    """
    from django.core.exceptions import ValidationError
    from django.db.models import Max
    from decimal import Decimal
    from .models import MovimientoKardex

    # Obtener todos los movimientos del lote
    movimientos_lote = MovimientoKardex.objects.filter(lote_carga=lote_id)
    if not movimientos_lote.exists():
        raise ValidationError(f"No se encontraron movimientos para el lote {lote_id}.")

    # Identificar artículos y almacenes afectados y su fecha máxima en el lote
    agrupado = movimientos_lote.values('articulo_id', 'almacen_id').annotate(
        max_fecha_lote=Max('fecha_hora')
    )

    articulos_comprometidos = []

    # 1. Validación Defensiva de Integridad Contable
    for group in agrupado:
        art_id = group['articulo_id']
        alm_id = group['almacen_id']
        max_fecha_lote = group['max_fecha_lote']

        # Buscar si hay algún movimiento de salida (Venta, Ajuste de salida, Consumo)
        # posterior a la fecha en que se cargó el lote.
        movimientos_posteriores = MovimientoKardex.objects.filter(
            articulo_id=art_id,
            almacen_id=alm_id,
            fecha_hora__gt=max_fecha_lote,
            tipo='SALIDA'
        )

        # Excluir de esta comprobación si hay movimientos que justamente
        # sean partes de revertir colisiones del mismo lote
        movimientos_posteriores = movimientos_posteriores.exclude(lote_carga=lote_id)

        if movimientos_posteriores.exists():
            articulo_obj = Articulo.objects.get(pk=art_id)
            articulos_comprometidos.append(articulo_obj.sku)

    if articulos_comprometidos:
        skus_str = ", ".join(articulos_comprometidos)
        raise ValidationError(
            f"El reverso fue bloqueado por seguridad contable. "
            f"Los siguientes SKU registran salidas posteriores a la carga masiva: {skus_str}"
        )

    # Calcular el neto ingresado por artículo y almacén
    # (ENTRADAS del lote - SALIDAS del lote en caso de SUSTITUIR)
    neto_ingresado = {}
    for mov in movimientos_lote:
        key = (mov.articulo_id, mov.almacen_id)
        if key not in neto_ingresado:
            neto_ingresado[key] = Decimal('0')
            
        if mov.tipo == 'ENTRADA':
            neto_ingresado[key] += mov.cantidad
        else:
            neto_ingresado[key] -= mov.cantidad

    # 2. Ejecución Atómica del Reverso (C-05: Conceptos estandarizados)
    detalle_reverso = f"Reverso automático de Carga Masiva - Lote {lote_id}"
    reversos_ejecutados = 0

    with transaction.atomic():
        for (art_id, alm_id), cantidad_neta in neto_ingresado.items():
            if cantidad_neta > 0:
                # Recuperar las instancias
                articulo = Articulo.objects.get(pk=art_id)
                almacen = Almacen.objects.get(pk=alm_id)
                
                # Efectuar una salida para descontar la cantidad neta ingresada
                registrar_movimiento(
                    articulo=articulo,
                    almacen=almacen,
                    tipo='SALIDA',
                    cantidad=cantidad_neta,
                    concepto='REVERSO_SALIDA',
                    detalle_adicional=detalle_reverso,
                    usuario=usuario
                )
                reversos_ejecutados += 1
            elif cantidad_neta < 0:
                # Si el neto fue negativo (caso extremadamente raro donde la carga
                # masiva quitó más de lo que puso, por ej. si Sustituir dejó un valor
                # menor al stock original), revertir la salida neta implicaría
                # devolver stock (ENTRADA). Para el lote, es una carga masiva.
                articulo = Articulo.objects.get(pk=art_id)
                almacen = Almacen.objects.get(pk=alm_id)
                registrar_movimiento(
                    articulo=articulo,
                    almacen=almacen,
                    tipo='ENTRADA',
                    cantidad=abs(cantidad_neta),
                    concepto='REVERSO_ENTRADA',
                    detalle_adicional=detalle_reverso,
                    usuario=usuario
                )
                reversos_ejecutados += 1

    logger.info("[REVERSO LOTE] Lote %s revertido exitosamente. %d registros afectados.", lote_id, reversos_ejecutados)
    
    return {
        'ok': True,
        'mensaje': f"Lote {lote_id} revertido con éxito. Kárdex actualizado.",
        'reversos_ejecutados': reversos_ejecutados
    }


# ─────────────────────────────────────────────────────────────────────────────
# 8. MÓDULO DE MOVIMIENTOS Y AJUSTES MANUALES
# ─────────────────────────────────────────────────────────────────────────────

def transferir_mercancia(articulo_sku: str, almacen_origen_id: int, almacen_destino_id: int, cantidad: Decimal, usuario: str = '') -> dict:
    """
    Transfiere stock físico de un almacén a otro de forma atómica y auditable.
    """
    from .models import Articulo, Almacen
    from .services import registrar_movimiento

    if almacen_origen_id == almacen_destino_id:
        raise ValueError("El almacén de origen y destino no pueden ser el mismo.")

    if cantidad <= 0:
        raise ValueError("La cantidad a transferir debe ser mayor a 0.")

    with transaction.atomic():
        articulo = Articulo.objects.get(sku=articulo_sku)
        origen = Almacen.objects.get(pk=almacen_origen_id)
        destino = Almacen.objects.get(pk=almacen_destino_id)

        # La validación de stock la hará automáticamente registrar_movimiento(..., 'SALIDA')
        # mediante get_stock_disponible() que implementa select_for_update() y
        # lanza ValueError si el stock es insuficiente.

        # 1. SALIDA del Origen (C-05: Concepto estandarizado)
        registrar_movimiento(
            articulo=articulo,
            almacen=origen,
            tipo='SALIDA',
            cantidad=cantidad,
            concepto='TRANSFERENCIA_SALIDA',
            detalle_adicional=f"Transferencia de salida hacia Almacén {destino.nombre}",
            usuario=usuario
        )

        # 2. ENTRADA al Destino (C-05: Concepto estandarizado)
        registrar_movimiento(
            articulo=articulo,
            almacen=destino,
            tipo='ENTRADA',
            cantidad=cantidad,
            concepto='TRANSFERENCIA_ENTRADA',
            detalle_adicional=f"Transferencia de entrada desde Almacén {origen.nombre}",
            usuario=usuario
        )

    logger.info("[TRANSFERENCIA] %s unid. de %s movidas de %s a %s.", cantidad, articulo_sku, origen.nombre, destino.nombre)
    return {'ok': True, 'mensaje': 'Transferencia ejecutada correctamente.'}


def ejecutar_ajuste_manual(articulo_sku: str, almacen_id: int, nueva_cantidad_fisica: Decimal, motivo: str, usuario: str = '') -> dict:
    """
    Asienta una diferencia matemática (Delta) para cuadrar el Kárdex
    con el inventario físico real.
    """
    from .models import Articulo, Almacen, InventarioAlmacen
    from .services import registrar_movimiento

    if nueva_cantidad_fisica < 0:
        raise ValueError("El inventario físico no puede ser negativo.")

    with transaction.atomic():
        articulo = Articulo.objects.get(sku=articulo_sku)
        almacen = Almacen.objects.get(pk=almacen_id)

        # Bloqueo pesimista del inventario
        try:
            inv = InventarioAlmacen.objects.select_for_update().get(
                articulo=articulo,
                almacen=almacen
            )
            stock_actual = inv.cantidad_disponible
        except InventarioAlmacen.DoesNotExist:
            stock_actual = Decimal('0')

        delta = nueva_cantidad_fisica - stock_actual

        if delta == 0:
            return {'ok': True, 'mensaje': 'El inventario ya concuerda. No se generaron movimientos.'}

        if delta > 0:
            registrar_movimiento(
                articulo=articulo,
                almacen=almacen,
                tipo='ENTRADA',
                cantidad=abs(delta),
                concepto='AJUSTE_ENTRADA',
                detalle_adicional=f"Ajuste manual de inventario (Suma): {motivo}",
                usuario=usuario
            )
        else:
            registrar_movimiento(
                articulo=articulo,
                almacen=almacen,
                tipo='SALIDA',
                cantidad=abs(delta),
                concepto='AJUSTE_SALIDA',
                detalle_adicional=f"Ajuste manual de inventario (Resta): {motivo}",
                usuario=usuario
            )

    logger.info("[AJUSTE MANUAL] %s en %s: Delta %s. Motivo: %s", articulo_sku, almacen.nombre, delta, motivo)
    return {'ok': True, 'mensaje': f"Ajuste realizado. Delta aplicado: {delta}"}


# ─────────────────────────────────────────────────────────────────────────────
# 9. CONTROL DE COSTOS Y COMPRAS
# ─────────────────────────────────────────────────────────────────────────────

def registrar_compra_proveedor(
    empresa_id=None, proveedor_id=None, tipo_documento='FACTURA_COMPRA',
    numero_factura='', fecha_compra=None, descuento_global=Decimal('0.00'),
    lista_items=None, almacen_id=None, usuario='',
    observaciones='', **kwargs) -> dict:
    """
    Registra el ingreso de mercancía por compra a proveedor mediante un Documento de Compra.
    Tipos: FACTURA_COMPRA, NOTA_ENTREGA_PROVEEDOR, REGISTRO_MENOR.
    Parámetros:
      - tipo_documento: 'FACTURA_COMPRA' | 'NOTA_ENTREGA_PROVEEDOR' | 'REGISTRO_MENOR'
      - numero_factura: obligatorio para FACTURA_COMPRA. Opcional para los demás.
        Único por empresa cuando no es vacío.
      - descuento_global: Decimal 0-100 (default 0)
      - lista_items: [{sku, cantidad, costo_factura, descuento_aplicado(optional), iva_porcentaje(optional), seriales(optional)}]
    Flujo atómico:
    1. Valida tenant, proveedor, almacén, items.
    2. Lee snapshot tasas de ConfiguracionEmpresa.
    3. Crea DocumentoCompra (con correlativo interno auto).
    4. Para cada item: crea DetalleDocumentoCompra (4 costos snapshot + IVA + descuento),
       actualiza costo base del artículo, recalcula precio venta, registra Kárdex ENTRADA,
       crea seriales si aplica.
    5. Calcula totales del documento (subtotal, descuento, IVA, total Bs) y persiste.
    """
    # Compatibilidad hacia atrás: si se pasa monto_total_usd (firma antigua), 
    # se ignora y se calcula desde los items.
    if 'monto_total_usd' in kwargs:
        import warnings
        warnings.warn(
            "El parámetro 'monto_total_usd' está deprecado. "
            "El total se calcula automáticamente desde los items.",
            DeprecationWarning, stacklevel=2
        )
    from .managers import get_current_empresa
    from .models import (
        Articulo, Almacen, Contacto, ConfiguracionEmpresa,
        DocumentoCompra, DetalleDocumentoCompra
    )
    from decimal import Decimal
    # Validación perimetral Multi-Tenant contra contaminación de sesión cruzada
    _ctx_empresa = get_current_empresa()

    if _ctx_empresa is None:
        raise ValueError("Seguridad Contable: No se detectó un contexto de Tenant activo para esta transacción.")

    if empresa_id is None:
        empresa_id = _ctx_empresa

    if not empresa_id:
        raise ValueError("Seguridad Contable: El identificador de la empresa emisora es obligatorio en el payload.")

    try:
        empresa_id_int = int(empresa_id)
        ctx_int = int(_ctx_empresa)
    except (ValueError, TypeError):
        raise ValueError("Seguridad Contable: El identificador del Tenant en el payload es inválido o ha sido alterado.")

    if empresa_id_int != ctx_int:
        raise ValueError("El contexto de la empresa ha cambiado en otra pestaña. Petición abortada por seguridad contable.")

    # Validaciones de tipos y rangos
    if tipo_documento not in ('FACTURA_COMPRA', 'NOTA_ENTREGA_PROVEEDOR', 'REGISTRO_MENOR'):
        raise ValueError(
            f"Tipo de documento inválido: '{tipo_documento}'. "
            "Debe ser 'FACTURA_COMPRA', 'NOTA_ENTREGA_PROVEEDOR' o 'REGISTRO_MENOR'."
        )
    if tipo_documento == 'FACTURA_COMPRA':
        if not numero_factura or not str(numero_factura).strip():
            raise ValueError("Seguridad Contable: numero_factura es obligatorio cuando tipo_documento='FACTURA_COMPRA'.")
        numero_factura = str(numero_factura).strip()
    else:
        # NOTA_ENTREGA_PROVEEDOR y REGISTRO_MENOR: numero_documento opcional
        numero_factura = str(numero_factura).strip() if numero_factura else ''

    try:
        descuento_global = Decimal(str(descuento_global))
    except (ValueError, TypeError):
        raise ValueError("descuento_global debe ser un número decimal entre 0 y 100.")
    if not (Decimal('0') <= descuento_global <= Decimal('100')):
        raise ValueError("descuento_global debe estar entre 0 y 100.")

    if not lista_items:
        raise ValueError("La lista de artículos no puede estar vacía.")

    with transaction.atomic():
        # Validación multi-tenant del almacén
        try:
            almacen = Almacen.objects.get(pk=almacen_id, empresa_id=empresa_id_int)
        except Almacen.DoesNotExist:
            raise ValueError(f"El almacen {almacen_id} no pertenece a la empresa activa.")

        # Validación multi-tenant del proveedor
        try:
            proveedor = Contacto.objects.get(
                pk=proveedor_id, tipo='PROVEEDOR', empresa_id=empresa_id_int
            )
        except Contacto.DoesNotExist:
            raise ValueError(
                f"El proveedor {proveedor_id} no pertenece a la empresa activa "
                f"o no existe."
            )

        # Validación fecha_compra (no futura)
        if fecha_compra:
            from datetime import date
            if isinstance(fecha_compra, str):
                fecha_compra = date.fromisoformat(fecha_compra)
            if fecha_compra > date.today():
                raise ValueError("La fecha de compra no puede ser futura.")

        # Snapshot de tasas
        config_tasa = ConfiguracionEmpresa.objects.get(empresa_id=empresa_id_int)
        tasa_bcv_snap = config_tasa.tasa_bcv
        tasa_mercado_snap = config_tasa.tasa_mercado
        factor_snap = config_tasa.factor_cobertura

        # Validar unicidad de numero_factura cuando viene informado
        # (FACTURA_COMPRA siempre trae valor obligatorio; NOTA_ENTREGA_PROVEEDOR
        #  puede traer el # de recibo del proveedor si se desea registrar).
        # REGISTRO_MENOR normalmente va sin número.
        if numero_factura:
            if DocumentoCompra.global_objects.filter(
                empresa_id=empresa_id_int,
                numero_factura=numero_factura
            ).exists():
                raise ValueError(
                    f"El número de documento '{numero_factura}' ya existe "
                    "para esta empresa. No se pueden duplicar documentos."
                )

        # Crear cabecera (los totales se recalculan abajo).
        # numero y numero_interno NO se pasan -> DocumentoCompra.save()
        # los auto-genera con Max('numero') + correlativo_inicial_factura_compra.
        documento = DocumentoCompra.objects.create(
            empresa_id=empresa_id_int,
            proveedor=proveedor,
            tipo_documento=tipo_documento,
            numero_factura=numero_factura,
            fecha_compra=fecha_compra,
            monto_total_usd=Decimal('0.0000'),  # se recalcula
            descuento_global=descuento_global,
            iva_total=Decimal('0.0000'),  # se recalcula
            tasa_bcv_aplicada=tasa_bcv_snap,
            tasa_mercado_aplicada=tasa_mercado_snap,
            factor_cobertura_aplicado=factor_snap,
            fuente_tasa='MANUAL',
            monto_total_bs_snapshot=Decimal('0.00'),
            observaciones=observaciones,
        )

        # Procesar ítems
        subtotal_usd = Decimal('0.0000')
        iva_total_usd = Decimal('0.0000')

        for item in lista_items:
            sku = item.get('sku') or item.get('articulo_sku')
            cantidad = Decimal(str(item.get('cantidad', 0)))
            costo_factura = Decimal(str(item.get('costo_factura', item.get('costo_base', 0))))
            descuento_aplicado = Decimal(str(item.get('descuento_aplicado', '0')))
            iva_porcentaje = Decimal(str(item.get('iva_porcentaje', '16.00')))

            if cantidad <= 0:
                raise ValueError(f"La cantidad comprada para {sku} debe ser mayor a 0.")
            if costo_factura < 0:
                raise ValueError(f"El costo para {sku} no puede ser negativo.")
            if not (Decimal('0') <= descuento_aplicado <= Decimal('100')):
                raise ValueError(f"descuento_aplicado para {sku} debe estar entre 0 y 100.")
            if not (Decimal('0') <= iva_porcentaje <= Decimal('100')):
                raise ValueError(f"iva_porcentaje para {sku} debe estar entre 0 y 100.")

            # Validación multi-tenant del artículo
            try:
                articulo = Articulo.objects.select_for_update().get(
                    sku=sku, empresa_id=empresa_id_int
                )
            except Articulo.DoesNotExist:
                raise ValueError(f"El articulo {sku} no pertence a la empresa activa.")

            # 4 snapshots de costos
            costo_ajustado = (costo_factura * factor_snap).quantize(Decimal('0.0001'))
            costo_directo_bcv = (costo_factura * tasa_bcv_snap).quantize(Decimal('0.01'))
            costo_ajustado_bcv = (costo_ajustado * tasa_bcv_snap).quantize(Decimal('0.01'))

            # Crear Detalle (snapshot inmutable)
            detalle = DetalleDocumentoCompra.objects.create(
                documento_compra=documento,
                articulo=articulo,
                almacen=almacen,
                cantidad=cantidad,
                costo_base=costo_factura,
                costo_ajustado=costo_ajustado,
                costo_directo_bcv=costo_directo_bcv,
                costo_ajustado_bcv=costo_ajustado_bcv,
                costo_unitario_snapshot=articulo.costo,
                descuento_aplicado=descuento_aplicado,
                iva_porcentaje=iva_porcentaje,
            )

            # Acumular totales del documento (usando propiedades del detalle)
            subtotal_usd += detalle.subtotal_usd
            iva_total_usd += detalle.iva_usd

            # Actualización de Costo Base + Recalculo Precio Venta
            articulo.costo = costo_factura
            margen = articulo.margen_ind
            if not margen or margen <= 0:
                config = ConfiguracionEmpresa.objects.get(empresa=articulo.empresa)
                margen = config.margen_global if config.margen_global else Decimal('30.00')

            metodo = (articulo.metodo_ganancia or 'MARKUP').upper()
            if metodo == 'MARKUP':
                precio_recalculado = costo_factura * (Decimal('1') + (margen / Decimal('100')))
            else:  # MARGIN
                factor_margen = Decimal('1') - (margen / Decimal('100'))
                if factor_margen <= 0:
                    raise ValueError("El margen comercial no puede ser igual o mayor al 100%.")
                precio_recalculado = costo_factura / factor_margen

            articulo.precio_divisa = precio_recalculado.quantize(Decimal('0.01'))
            articulo.save(update_fields=['costo', 'precio_divisa', 'fecha_actualizacion'])

            # Kárdex ENTRADA
            registrar_movimiento(
                articulo=articulo,
                almacen=almacen,
                tipo='ENTRADA',
                cantidad=cantidad,
                concepto='COMPRA',
                usuario=usuario,
                documento_compra=documento,
            )

            # Seriales (si usa_serial y vienen en item)
            seriales = item.get('seriales', [])
            if articulo.usa_serial:
                if not seriales:
                    raise ValueError(f"El artículo '{articulo.nombre}' requiere seriales, pero no se enviaron.")
                from .models import SerialArticulo
                seriales_unicos = list(set(seriales))
                if len(seriales_unicos) != int(cantidad):
                    raise ValueError(
                        f"El artículo '{articulo.nombre}' requiere exactamente {int(cantidad)} seriales, "
                        f"pero se enviaron {len(seriales_unicos)}."
                    )
                nuevos_seriales = [
                    SerialArticulo(
                        empresa=articulo.empresa,
                        articulo=articulo,
                        serial=serial,
                        almacen=almacen,
                        estado='DISPONIBLE',
                        compra_origen=documento
                    ) for serial in seriales_unicos
                ]
                SerialArticulo.objects.bulk_create(nuevos_seriales)

            logger.info("[COMPRA] %s unid. de %s. Costo actualizado a %s", cantidad, sku, costo_factura)

        # Totales finales del documento
        # Aplicar descuento global proporcionalmente a cada ítem, luego calcular IVA
        factor_global = Decimal('1') - (descuento_global / Decimal('100'))
        subtotal_after_discount = Decimal('0')
        iva_total_usd = Decimal('0')
        
        for d in documento.detalles.all():
            item_subtotal = d.subtotal_usd
            item_discounted = (item_subtotal * factor_global).quantize(Decimal('0.0001'))
            item_iva = (item_discounted * (d.iva_porcentaje / Decimal('100'))).quantize(Decimal('0.0001'))
            subtotal_after_discount += item_discounted
            iva_total_usd += item_iva

        monto_total_usd = (subtotal_after_discount + iva_total_usd).quantize(Decimal('0.0001'))

        # Total en Bs = (subtotal_with_discount + iva_total_usd) * tasa_bcv.
        # El IVA en Bs se calcula dinámicamente en la property iva_bs de cada
        # DetalleDocumentoCompra (iva_usd * tasa_bcv_aplicada) y se muestra
        # en vistas/PDF. Aquí calculamos el snapshot monto_total_bs_snapshot
        # del documento completo aplicando el descuento_global a la base
        # monetaria en USD antes de convertir a Bs.
        total_bs = ((subtotal_after_discount + iva_total_usd) * tasa_bcv_snap).quantize(Decimal('0.01'))

        # Actualizar documento con totales finales
        documento.monto_total_usd = monto_total_usd
        documento.iva_total = iva_total_usd
        documento.monto_total_bs_snapshot = total_bs
        documento.save(update_fields=['monto_total_usd', 'iva_total', 'monto_total_bs_snapshot'])

        return {
            'ok': True,
            'mensaje': f"Compra {documento.numero_interno} registrada exitosamente con {len(lista_items)} ítems.",
            'documento_id': documento.id,
            'numero_interno': documento.numero_interno,
            'numero_factura': documento.numero_factura,
        }


# ─────────────────────────────────────────────────────────────────────────────
# 10. EXPORTACIÓN LÓGICA POR TENANT (TICKET #13)
# ─────────────────────────────────────────────────────────────────────────────

def exportar_datos_tenant(empresa_id: int, meses_historico: int = 6) -> dict:
    """
    Realiza una extracción lógica estructurada aislando la data por Tenant.
    Truncando el historial de Kárdex y Notas de Entrega para evitar timeouts HTTP.
    Retorna un diccionario de Python serializable a JSON.
    """
    from datetime import timedelta
    from django.utils import timezone
    from .models import Articulo, Almacen, Contacto, NotaEntrega, MovimientoKardex, Empresa

    empresa = Empresa.objects.get(pk=empresa_id)
    fecha_limite = timezone.now() - timedelta(days=30 * meses_historico)

    # 1. Artículos
    articulos = list(Articulo.global_objects.filter(empresa=empresa).values())
    
    # 2. Almacenes
    almacenes = list(Almacen.global_objects.filter(empresa=empresa).values())
    
    # 3. Contactos
    contactos = list(Contacto.global_objects.filter(empresa=empresa).values())
    
    # 4. Notas de Entrega (histórico mitigado)
    notas = list(NotaEntrega.global_objects.filter(
        empresa=empresa, 
        fecha__gte=fecha_limite
    ).values())
    
    # 5. Kárdex (histórico mitigado)
    kardex = list(MovimientoKardex.global_objects.filter(
        articulo__empresa=empresa, 
        fecha_hora__gte=fecha_limite
    ).values())

    payload = {
        'metadata': {
            'empresa_id': empresa.pk,
            'empresa_nombre': empresa.nombre,
            'empresa_rif': empresa.rif,
            'fecha_exportacion': timezone.now().isoformat(),
            'meses_historico': meses_historico
        },
        'data': {
            'articulos': articulos,
            'almacenes': almacenes,
            'contactos': contactos,
            'notas_entrega': notas,
            'movimientos_kardex': kardex
        }
    }

    # y devolvemos un dict parseado, o el string según se prefiera.
    # La firma dice "dict" en la instrucción, devolvemos el dict limpio,
    # y en la vista haremos json.dumps() con el DjangoJSONEncoder.
    return payload


# ─────────────────────────────────────────────────────────────────────────────
# 12. MÓDULO DE CONTRAPARTIDAS Y REVERSOS (TICKET #20)
# ─────────────────────────────────────────────────────────────────────────────
#
# Nota histórica: La versión original de "MÓDULO DE DEVOLUCIONES Y CUARENTENA"
# (Ticket #15-SAAS, MVP antiguo) tenía una función `procesar_devolucion_venta`
# con firma posicional y soporte para "tipo_costo" (HISTORICO|ACTUAL),
# "es_defectuoso" (merma automática) y "almacen_cuarentena". Esa API fue
# enteramente reemplazada en el Ticket #18-NC (ADR-29) por un diseño limpio
# 1-NC-1-origen con snapshots de precio e IVA persistidos por línea. La
# maquinaria nueva vive en la sección 14 (más abajo), y los tests legacy en
# TestNotasDeCreditoPOS están marcados con `@skip(reason=...)` por compatibilidad
# documental.


@transaction.atomic
def reversar_nota_entrega(empresa_id: int, nota_id: int, motivo: str) -> dict:
    from .models import NotaEntrega, SerialArticulo
    from .services import registrar_movimiento

    nota = NotaEntrega.objects.select_for_update().get(id=nota_id, empresa_id=empresa_id)
    if nota.estado == 'ANULADO':
        raise ValueError("La Nota de Entrega ya se encuentra anulada.")
    
    nota.estado = 'ANULADO'
    nota.motivo_anulacion = motivo
    nota.save(update_fields=['estado', 'motivo_anulacion'])
    
    # Revertir detalles
    for detalle in nota.detalles.all():
        registrar_movimiento(
            articulo=detalle.articulo,
            almacen=nota.almacen,
            tipo='ENTRADA',
            cantidad=detalle.cantidad,
            concepto='DEVOLUCION_VENTA',
            detalle_adicional=f"Reverso de Venta (Nota #{nota.numero}): {motivo}",
            nota_entrega=nota,
            usuario='Sistema'
        )
        
        # Liberar seriales vinculados
        seriales = SerialArticulo.objects.filter(detalle_nota=detalle, empresa_id=empresa_id)
        seriales.update(estado='DISPONIBLE', detalle_nota=None)
        
    return {'ok': True, 'mensaje': f"Nota de Entrega #{nota.numero} reversada exitosamente."}


@transaction.atomic
def reversar_documento_compra(empresa_id: int, compra_id: int, motivo: str) -> dict:
    from .models import DocumentoCompra, SerialArticulo, MovimientoKardex
    from .services import registrar_movimiento

    compra = DocumentoCompra.objects.select_for_update().get(id=compra_id, empresa_id=empresa_id)
    if compra.estado == 'ANULADO':
        raise ValueError("El Documento de Compra ya se encuentra anulado.")
    
    compra.estado = 'ANULADO'
    compra.motivo_anulacion = motivo
    compra.save(update_fields=['estado', 'motivo_anulacion'])
    
    # Para reversar la compra, buscamos los movimientos de Kárdex asociados a ella
    # que fueron generados durante la compra ('ENTRADA' por concepto 'COMPRA')
    movimientos_compra = MovimientoKardex.objects.filter(documento_compra=compra, tipo='ENTRADA', concepto='COMPRA')
    
    for mov in movimientos_compra:
        registrar_movimiento(
            articulo=mov.articulo,
            almacen=mov.almacen,
            tipo='SALIDA',
            cantidad=mov.cantidad,
            concepto='ANULACION_COMPRA',
            detalle_adicional=f"Reverso de Compra (Factura {compra.numero_factura}): {motivo}",
            usuario='Sistema',
            documento_compra=compra
        )
    
    # Anular seriales provenientes de esta compra
    seriales = SerialArticulo.objects.filter(compra_origen=compra, empresa_id=empresa_id)
    seriales.update(estado='ANULADO_COMPRA')
    
    return {'ok': True, 'mensaje': f"Compra Factura {compra.numero_factura} reversada exitosamente."}


# ─────────────────────────────────────────────────────────────────────────────
# 14. NOTAS DE CRÉDITO — DEVOLUCIONES PARCIALES O TOTALES (TICKET #18-NC)
#
# Reglas (ADR-29):
#   - Una NC se apega a un único documento origen (NE en venta o
#     DocumentoCompra en compra), enforced por CheckConstraint.
#   - Múltiples NCs pueden existir sobre el mismo DetalleNotaEntrega /
#     DetalleDocumentoCompra; cada una respeta `cantidad_pendiente_devolver`.
#   - El kardex guarda Movimiento 'DEVOLUCION_VENTA' (entrada) o
#     'DEVOLUCION_COMPRA' (salida) por la cantidad_devuelta.
#   - Los seriales se liberan FIFO (estado=VENDIDO → DISPONIBLE) hasta
#     cubrir la cantidad_devuelta (solo en devoluciones de venta).
#   - Precios snapshot persisten incluso si Articulo cambia después.
# ─────────────────────────────────────────────────────────────────────────────


@transaction.atomic
def procesar_devolucion_venta(
    empresa_id: int,
    nota_id: int,
    items_devueltos: list,
    motivo: str,
    usuario: str = 'Sistema',
) -> dict:
    """
    Crea una NotaCredito de VENTA con N DetalleNotaCredito.
    
    Par\u00e1metros:
      - empresa_id: tenant activo (seguridad multi-tenant).
      - nota_id: pk de NotaEntrega (origen de la devoluci\u00f3n).
      - items_devueltos: lista de dicts {
            'detalle_id': int (pk de DetalleNotaEntrega),
            'cantidad_devolver': Decimal,
            'precio_unitario': Decimal (opcional; default = precio_base origen),
            'iva_porcentaje': Decimal (opcional; default = iva_porcentaje origen)
        }
      - motivo: requerido, no blank.
      - usuario: nombre del operador (default 'Sistema').
    
    Retorna: dict {
        'ok': True,
        'nc_id': ...,
        'numero_control': 'NC-00000001',
        'detalles_ids': [...],
        'monto_total_reembolso': Decimal,
        'mensaje': '...'
    }
    
    Lanza ValueError si:
      - motivo blank.
      - nota no encontrada / otro tenant.
      - detalle_id no pertenece a la nota.
      - cantidad_devolver > cantidad_pendiente_devolver del detalle.
      - precio_unitario < 0 o iva fuera de [0, 100].
      - items_devueltos vac\u00edo.
    """
    from decimal import Decimal as _D
    from .models import (
        NotaEntrega, DetalleNotaEntrega, NotaCredito, DetalleNotaCredito,
        SerialArticulo,
    )

    motivo = (motivo or '').strip()
    if not motivo:
        raise ValueError("El motivo de la NC es obligatorio.")
    if not items_devueltos:
        raise ValueError("La lista de items devueltos no puede estar vacía.")

    # 1. Validación perimetral multi-tenant (con select_for_update).
    try:
        nota = NotaEntrega.objects.select_for_update().get(
            pk=nota_id, empresa_id=empresa_id,
        )
    except NotaEntrega.DoesNotExist:
        raise ValueError(
            f"La Nota de Entrega {nota_id} no pertenece a la empresa activa."
        )

    if nota.estado == 'ANULADO':
        raise ValueError("La Nota de Entrega está anulada; no se puede emitir una NC sobre ella.")

    # 2. Crear cabecera (NotaCredito.save() calcula numero y numero_control).
    nc = NotaCredito.objects.create(
        empresa_id=empresa_id,
        nota_entrega=nota,
        doc_origen_tipo='VENTA',
        motivo=motivo,
        usuario=usuario,
        estado='PROCESADO',
    )

    detalles_ids = []
    monto_total_usd = _D('0.0000')

    try:
        for item in items_devueltos:
            cantidad_devolver = _D(str(item.get('cantidad_devolver') or 0))
            if cantidad_devolver <= 0:
                raise ValueError("cantidad_devolver debe ser > 0 (omite el item si 0).")

            try:
                detalle = DetalleNotaEntrega.objects.select_for_update().get(
                    pk=item['detalle_id'],
                    nota_entrega=nota,
                )
            except DetalleNotaEntrega.DoesNotExist:
                raise ValueError(
                    f"El detalle {item['detalle_id']} no pertenece a la nota {nota_id}."
                )

            # 3. Validar cantidad pendiente.
            cantidad_pendiente = detalle.cantidad_pendiente_devolver
            if cantidad_devolver > cantidad_pendiente:
                raise ValueError(
                    f"No se pueden devolver {cantidad_devolver} de '{detalle.articulo.nombre}'. "
                    f"Cantidad pendiente_devolver = {cantidad_pendiente}."
                )

            # 4. Snapshots (override opcional del precio / iva).
            precio_unitario = _D(str(
                item.get('precio_unitario', detalle.precio_base)
            ))
            iva_pct = _D(str(
                item.get('iva_porcentaje', detalle.iva_porcentaje or 0)
            ))
            if precio_unitario < 0:
                raise ValueError("precio_unitario no puede ser negativo.")
            if not (_D('0') <= iva_pct <= _D('100')):
                raise ValueError(
                    f"iva_porcentaje para {detalle.articulo.sku} debe estar entre 0 y 100."
                )

            # 5. Crear DetalleNotaCredito (save() calcula totales snapshot).
            det_nc = DetalleNotaCredito.objects.create(
                nota_credito=nc,
                detalle_origen_venta=detalle,
                articulo=detalle.articulo,
                almacen=detalle.almacen,
                cantidad_devuelta=cantidad_devolver,
                precio_unitario_snapshot=precio_unitario,
                iva_porcentaje_snapshot=iva_pct,
            )
            detalles_ids.append(det_nc.pk)
            monto_total_usd += det_nc.linea_total_usd

            # 6. Movimiento de Kárdex: ENTRADA por la cantidad devuelta
            #    (reingresa stock al almacén del doc. origen).
            registrar_movimiento(
                articulo=detalle.articulo,
                almacen=detalle.almacen,
                tipo='ENTRADA',
                cantidad=cantidad_devolver,
                concepto='DEVOLUCION_VENTA',
                detalle_adicional=(
                    f"NC {nc.numero_control} sobre NotEntrega NE-{nota.numero:08d}: {motivo}"
                ),
                nota_entrega=nota,
                usuario=usuario,
            )

            # 7. Liberar seriales FIFO (estado VENDIDO → DISPONIBLE).
            if hasattr(detalle.articulo, 'usa_serial') and detalle.articulo.usa_serial:
                seriales_disponibles = list(
                    SerialArticulo.objects.filter(
                        detalle_nota=detalle,
                        estado='VENDIDO',
                        empresa_id=empresa_id,
                    ).order_by('id')[:int(cantidad_devolver)]
                )
                if len(seriales_disponibles) < int(cantidad_devolver):
                    # Log warning pero no abortamos: el operador puede seguir
                    # si el producto tiene stock previo sin seriales.
                    import logging
                    logging.warning(
                        "NC %s: seriales insuficientes para '%s' "
                        "(esperados=%s, encontrados=%s). Continuando sin liberar todos.",
                        nc.numero_control,
                        detalle.articulo.sku,
                        int(cantidad_devolver),
                        len(seriales_disponibles),
                    )
                SerialArticulo.objects.filter(
                    pk__in=[s.pk for s in seriales_disponibles]
                ).update(estado='DISPONIBLE', detalle_nota=None)
    except Exception:
        # Si algo falla, el @transaction.atomic envuelve todo y hace rollback.
        nc.delete() if nc.pk else None
        raise

    nc.monto_total_reembolso = monto_total_usd.quantize(_D('0.0001'))
    nc.save(update_fields=['monto_total_reembolso'])

    return {
        'ok': True,
        'nc_id': nc.pk,
        'numero_control': nc.numero_control,
        'detalles_ids': detalles_ids,
        'monto_total_reembolso': nc.monto_total_reembolso,
        'mensaje': (
            f"Nota de Crédito {nc.numero_control} emitida sobre "
            f"Nota de Entrega NE-{nota.numero:08d}: {len(detalles_ids)} item(s) devuelto(s)."
        ),
    }


@transaction.atomic
def procesar_devolucion_compra(
    empresa_id: int,
    compra_id: int,
    items_devueltos: list,
    motivo: str,
    usuario: str = 'Sistema',
) -> dict:
    """
    Crea una NotaCredito de COMPRA con N DetalleNotaCredito.
    
    An\u00e1logo a procesar_devolucion_venta pero:
      - kardex: SALIDA por la cantidad (sacamos del almacén del doc. origen).
      - seriales se reciclan a estado ANULADO_COMPRA (no DISPONIBLE).
    
    Mismas validaciones que procesar_devolucion_venta.
    """
    from decimal import Decimal as _D
    from .models import (
        DocumentoCompra, DetalleDocumentoCompra, NotaCredito, DetalleNotaCredito,
    )

    motivo = (motivo or '').strip()
    if not motivo:
        raise ValueError("El motivo de la NC es obligatorio.")
    if not items_devueltos:
        raise ValueError("La lista de items devueltos no puede estar vacía.")

    try:
        compra = DocumentoCompra.objects.select_for_update().get(
            pk=compra_id, empresa_id=empresa_id,
        )
    except DocumentoCompra.DoesNotExist:
        raise ValueError(
            f"El Documento de Compra {compra_id} no pertenece a la empresa activa."
        )

    if compra.estado == 'ANULADO':
        raise ValueError("La Compra está anulada; no se puede emitir una NC sobre ella.")

    nc = NotaCredito.objects.create(
        empresa_id=empresa_id,
        factura_compra=compra,
        doc_origen_tipo='COMPRA',
        motivo=motivo,
        usuario=usuario,
        estado='PROCESADO',
    )

    detalles_ids = []
    monto_total_usd = _D('0.0000')

    try:
        for item in items_devueltos:
            cantidad_devolver = _D(str(item.get('cantidad_devolver') or 0))
            if cantidad_devolver <= 0:
                raise ValueError("cantidad_devolver debe ser > 0 (omite el item si 0).")

            try:
                detalle = DetalleDocumentoCompra.objects.select_for_update().get(
                    pk=item['detalle_id'],
                    documento_compra=compra,
                )
            except DetalleDocumentoCompra.DoesNotExist:
                raise ValueError(
                    f"El detalle de compra {item['detalle_id']} no pertenece al doc. {compra_id}."
                )

            cantidad_pendiente = detalle.cantidad_pendiente_devolver
            if cantidad_devolver > cantidad_pendiente:
                raise ValueError(
                    f"No se pueden devolver {cantidad_devolver} de '{detalle.articulo.nombre}'. "
                    f"Cantidad pendiente_devolver = {cantidad_pendiente}."
                )

            precio_unitario = _D(str(
                item.get('precio_unitario', detalle.costo_base)
            ))
            iva_pct = _D(str(
                item.get('iva_porcentaje', detalle.iva_porcentaje or 0)
            ))
            if precio_unitario < 0:
                raise ValueError("precio_unitario no puede ser negativo.")
            if not (_D('0') <= iva_pct <= _D('100')):
                raise ValueError(
                    f"iva_porcentaje para {detalle.articulo.sku} debe estar entre 0 y 100."
                )

            det_nc = DetalleNotaCredito.objects.create(
                nota_credito=nc,
                detalle_origen_compra=detalle,
                articulo=detalle.articulo,
                almacen=detalle.almacen,
                cantidad_devuelta=cantidad_devolver,
                precio_unitario_snapshot=precio_unitario,
                iva_porcentaje_snapshot=iva_pct,
            )
            detalles_ids.append(det_nc.pk)
            monto_total_usd += det_nc.linea_total_usd

            # Kardex: SALIDA porque estamos sacando del almacén por devolución.
            registrar_movimiento(
                articulo=detalle.articulo,
                almacen=detalle.almacen,
                tipo='SALIDA',
                cantidad=cantidad_devolver,
                concepto='DEVOLUCION_COMPRA',
                detalle_adicional=(
                    f"NC {nc.numero_control} sobre DocumentoCompra "
                    f"{compra.tipo_documento} {compra.numero_factura or compra.numero_interno}: {motivo}"
                ),
                usuario=usuario,
                documento_compra=compra,
            )

            # Seriales: ANULADO_COMPRA (no DISPONIBLE; ya no son stock).
            if hasattr(detalle.articulo, 'usa_serial') and detalle.articulo.usa_serial:
                from .models import SerialArticulo
                seriales = list(
                    SerialArticulo.objects.filter(
                        compra_origen=compra,
                        articulo=detalle.articulo,
                        estado='DISPONIBLE',
                        empresa_id=empresa_id,
                    ).order_by('id')[:int(cantidad_devolver)]
                )
                if seriales:
                    SerialArticulo.objects.filter(
                        pk__in=[s.pk for s in seriales]
                    ).update(estado='ANULADO_COMPRA')
    except Exception:
        nc.delete() if nc.pk else None
        raise

    nc.monto_total_reembolso = monto_total_usd.quantize(_D('0.0001'))
    nc.save(update_fields=['monto_total_reembolso'])

    return {
        'ok': True,
        'nc_id': nc.pk,
        'numero_control': nc.numero_control,
        'detalles_ids': detalles_ids,
        'monto_total_reembolso': nc.monto_total_reembolso,
        'mensaje': (
            f"Nota de Crédito {nc.numero_control} emitida sobre Compra "
            f"#{compra_id}: {len(detalles_ids)} item(s) devuelto(s)."
        ),
    }

